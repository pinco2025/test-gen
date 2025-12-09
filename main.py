import yaml
import openpyxl
from openpyxl import Workbook, load_workbook
import sqlite3
import hashlib
import os
import sys
import argparse
import json
from pathlib import Path
from datetime import datetime

class QuestionProcessor:
    def __init__(self, excel_path='questions_database.xlsx', db_path='questions_database.db', max_tags=4, auto_init=True):
        self.excel_path = excel_path
        self.db_path = db_path
        self.max_tags = max_tags
        self.headers = [
            'UUID', 'Question', 'Question_Image_URL',
            'Option_A', 'Option_A_Image_URL',
            'Option_B', 'Option_B_Image_URL',
            'Option_C', 'Option_C_Image_URL',
            'Option_D', 'Option_D_Image_URL',
            'Answer', 'Type', 'Year'
        ] + [f'Tag_{i}' for i in range(1, max_tags + 1)]
        
        # Initialize database only if auto_init is True
        if auto_init:
            self.init_database()
    
    def init_database(self):
        """Initialize SQLite database with required tables"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Create questions table
        tag_columns = ', '.join([f'tag_{i} TEXT' for i in range(1, self.max_tags + 1)])
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS questions (
                uuid TEXT PRIMARY KEY,
                question TEXT NOT NULL,
                question_image_url TEXT,
                option_a TEXT,
                option_a_image_url TEXT,
                option_b TEXT,
                option_b_image_url TEXT,
                option_c TEXT,
                option_c_image_url TEXT,
                option_d TEXT,
                option_d_image_url TEXT,
                answer TEXT NOT NULL,
                type TEXT NOT NULL,
                year TEXT,
                {tag_columns},
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create indexes for faster searches
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_question_text 
            ON questions(question)
        ''')
        
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_type 
            ON questions(type)
        ''')
        
        conn.commit()
        conn.close()
        
        if not os.path.exists(self.db_path) or os.path.getsize(self.db_path) == 0:
            print(f"✓ Created new SQLite database: {self.db_path}")
        
    def generate_uuid(self, question_text, type_val, year):
        """Generate UUID using hash of question + metadata"""
        combined = f"{question_text}{type_val}{year}"
        hash_obj = hashlib.md5(combined.encode())
        return hash_obj.hexdigest()[:12]  # 12 character hash
    
    def load_yaml(self, yaml_path):
        """Load and parse YAML file"""
        if not os.path.exists(yaml_path):
            raise FileNotFoundError(f"YAML file not found: {yaml_path}")
        
        with open(yaml_path, 'r', encoding='utf-8') as file:
            data = yaml.safe_load(file)
        
        # Validate required metadata
        if 'type' not in data:
            raise ValueError("Missing required metadata field: 'type'")
        
        # Validate questions exist
        if 'questions' not in data or not data['questions']:
            raise ValueError("No questions found in YAML file")
        
        return data
    
    def load_excel(self):
        """Load existing Excel file"""
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(
                f"Excel file not found: {self.excel_path}\n"
                f"Please create the Excel file first or check the path."
            )
        
        return load_workbook(self.excel_path)
    
    def create_excel_template(self):
        """Create a new Excel template file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Questions"
        
        # Write headers
        for col, header in enumerate(self.headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(self.excel_path)
        print(f"✓ Created new Excel template: {self.excel_path}")
    
    def get_existing_questions_excel(self, ws):
        """Get all existing question texts from Excel"""
        existing = set()
        for row in range(2, ws.max_row + 1):
            question = ws.cell(row=row, column=2).value  # Column 2 is Question
            if question:
                existing.add(question.strip().lower())
        return existing
    
    def get_existing_questions_db(self):
        """Get all existing question texts from SQLite database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT question FROM questions")
        existing = {row[0].strip().lower() for row in cursor.fetchall()}
        conn.close()
        return existing
    
    def parse_option(self, option_data):
        """Parse option data from YAML - returns (text, image_url)"""
        if not option_data:
            return '', ''
        
        if isinstance(option_data, dict):
            text = option_data.get('text', '')
            image_url = option_data.get('image_url', '')
            # Option can have just image URL without text
            return text, image_url
        else:
            # If option is just a string
            return str(option_data), ''
    
    def insert_question_db(self, question_data):
        """Insert a question into SQLite database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Prepare column names and values
        columns = [
            'uuid', 'question', 'question_image_url',
            'option_a', 'option_a_image_url',
            'option_b', 'option_b_image_url',
            'option_c', 'option_c_image_url',
            'option_d', 'option_d_image_url',
            'answer', 'type', 'year'
        ]
        columns.extend([f'tag_{i}' for i in range(1, self.max_tags + 1)])
        
        placeholders = ', '.join(['?' for _ in columns])
        columns_str = ', '.join(columns)
        
        try:
            cursor.execute(
                f"INSERT INTO questions ({columns_str}) VALUES ({placeholders})",
                question_data
            )
            conn.commit()
            return True
        except sqlite3.IntegrityError:
            # Duplicate UUID
            return False
        finally:
            conn.close()
    
    def process_single_file(self, yaml_path, skip_duplicates=False, overwrite_duplicates=False):
        """Process a single YAML file and append to Excel and SQLite"""
        print(f"\n{'='*60}")
        print(f"Processing: {yaml_path}")
        print(f"{'='*60}\n")
        
        # Load YAML data
        try:
            data = self.load_yaml(yaml_path)
        except Exception as e:
            print(f"✗ Error loading YAML: {e}")
            return False
        
        # Load Excel
        try:
            wb = self.load_excel()
            ws = wb.active
        except Exception as e:
            print(f"✗ Error loading Excel: {e}")
            return False
        
        # Get existing questions from both sources
        existing_questions_excel = self.get_existing_questions_excel(ws)
        existing_questions_db = self.get_existing_questions_db()
        existing_questions = existing_questions_excel | existing_questions_db
        
        # Extract metadata
        metadata = {
            'type': data.get('type', ''),
            'year': data.get('year', '')
        }
        
        print(f"Metadata:")
        print(f"  Type: {metadata['type']}")
        print(f"  Year: {metadata['year'] if metadata['year'] else 'Not specified'}")
        print(f"\nTotal questions in YAML: {len(data['questions'])}\n")
        
        # Find next empty row in Excel
        next_row = ws.max_row + 1
        if ws.cell(row=2, column=1).value is None:  # If first data row is empty
            next_row = 2
        
        # Process questions
        added_count = 0
        skipped_count = 0
        duplicate_questions = []
        
        for idx, q in enumerate(data['questions'], start=1):
            question_text = q.get('question', '').strip()
            
            # Check for duplicates
            if question_text.lower() in existing_questions:
                duplicate_questions.append((idx, question_text))
                skipped_count += 1
                continue
            
            # Extract question image URL first (needed for validation)
            question_image_url = q.get('image_url', '')
            
            # Validate question structure - must have either text or image
            if not question_text and not question_image_url:
                print(f"⚠ Warning: Question {idx} has neither text nor image. Skipping.")
                skipped_count += 1
                continue
            
            required_fields = ['A', 'B', 'C', 'D', 'answer']
            missing_fields = [f for f in required_fields if f not in q]
            if missing_fields:
                print(f"⚠ Warning: Question {idx} missing fields {missing_fields}. Skipping.")
                skipped_count += 1
                continue
            
            # Generate UUID
            uuid = self.generate_uuid(
                question_text,
                metadata['type'],
                metadata['year']
            )
            
            # Extract tags
            tags = q.get('tags', [])
            if len(tags) > self.max_tags:
                print(f"⚠ Warning: Question {idx} has {len(tags)} tags. Using first {self.max_tags}.")
                tags = tags[:self.max_tags]
            
            # Parse options
            option_a_text, option_a_image = self.parse_option(q.get('A'))
            option_b_text, option_b_image = self.parse_option(q.get('B'))
            option_c_text, option_c_image = self.parse_option(q.get('C'))
            option_d_text, option_d_image = self.parse_option(q.get('D'))
            
            # Validate that each option has either text or image
            for opt_label, opt_text, opt_image in [
                ('A', option_a_text, option_a_image),
                ('B', option_b_text, option_b_image),
                ('C', option_c_text, option_c_image),
                ('D', option_d_text, option_d_image)
            ]:
                if not opt_text and not opt_image:
                    print(f"⚠ Warning: Question {idx}, Option {opt_label} has neither text nor image. Skipping question.")
                    skipped_count += 1
                    continue
            
            # Prepare row data
            row_data = [
                uuid,
                question_text,
                question_image_url,
                option_a_text,
                option_a_image,
                option_b_text,
                option_b_image,
                option_c_text,
                option_c_image,
                option_d_text,
                option_d_image,
                q.get('answer', ''),
                metadata['type'],
                metadata['year']
            ]
            
            # Add tags
            for i in range(self.max_tags):
                row_data.append(tags[i] if i < len(tags) else '')
            
            # Write to Excel
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=next_row, column=col, value=value)
            
            # Write to SQLite
            if self.insert_question_db(row_data):
                next_row += 1
                added_count += 1
                existing_questions.add(question_text.lower())
            else:
                print(f"⚠ Warning: Failed to insert question {idx} into database.")
                skipped_count += 1
        
        # Handle duplicates
        if duplicate_questions and not skip_duplicates and not overwrite_duplicates:
            print(f"\n{'!'*60}")
            print(f"Found {len(duplicate_questions)} duplicate question(s):")
            for idx, q_text in duplicate_questions[:5]:  # Show first 5
                print(f"  • Question {idx}: {q_text[:60]}...")
            if len(duplicate_questions) > 5:
                print(f"  ... and {len(duplicate_questions) - 5} more")
            print(f"{'!'*60}\n")
            
            response = input("Skip duplicates? (y/n): ").strip().lower()
            if response != 'y':
                print("✗ Operation cancelled. No changes saved.")
                return False
        
        # Save Excel
        try:
            wb.save(self.excel_path)
            print(f"\n{'='*60}")
            print(f"✓ Successfully processed!")
            print(f"  Added: {added_count} questions")
            print(f"  Skipped: {skipped_count} questions")
            print(f"  Excel total: {ws.max_row - 1} questions")
            print(f"  SQLite total: {self.get_db_count()} questions")
            print(f"{'='*60}\n")
            return True
        except Exception as e:
            print(f"✗ Error saving Excel: {e}")
            return False
    
    def get_db_count(self):
        """Get total count of questions in SQLite database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM questions")
        count = cursor.fetchone()[0]
        conn.close()
        return count
    
    def process_batch(self, directory, skip_duplicates=False, overwrite_duplicates=False):
        """Process all YAML files in a directory"""
        yaml_files = list(Path(directory).glob('*.yaml')) + list(Path(directory).glob('*.yml'))
        
        if not yaml_files:
            print(f"✗ No YAML files found in: {directory}")
            return
        
        print(f"\nFound {len(yaml_files)} YAML file(s) to process\n")
        
        success_count = 0
        for yaml_file in yaml_files:
            if self.process_single_file(str(yaml_file), skip_duplicates, overwrite_duplicates):
                success_count += 1
        
        print(f"\n{'='*60}")
        print(f"Batch processing complete!")
        print(f"  Successfully processed: {success_count}/{len(yaml_files)} files")
        print(f"{'='*60}\n")
    
    def setup(self):
        """Initialize the question database system - create both Excel and SQLite databases"""
        print(f"\n{'='*60}")
        print(f"Question Database Setup")
        print(f"{'='*60}\n")
        
        setup_success = True
        
        # Create SQLite database
        print(f"[1/2] Initializing SQLite database...")
        try:
            self.init_database()
            if os.path.exists(self.db_path):
                print(f"      ✓ SQLite database ready: {self.db_path}")
            else:
                print(f"      ✗ Failed to create SQLite database")
                setup_success = False
        except Exception as e:
            print(f"      ✗ Error creating SQLite database: {e}")
            setup_success = False
        
        # Create Excel template
        print(f"\n[2/2] Creating Excel template...")
        try:
            if os.path.exists(self.excel_path):
                response = input(f"      Excel file already exists: {self.excel_path}\n      Overwrite? (y/n): ").strip().lower()
                if response != 'y':
                    print(f"      ⊙ Keeping existing Excel file")
                else:
                    self.create_excel_template()
            else:
                self.create_excel_template()
        except Exception as e:
            print(f"      ✗ Error creating Excel template: {e}")
            setup_success = False
        
        # Summary
        print(f"\n{'='*60}")
        if setup_success:
            print(f"✓ Setup Complete!")
            print(f"\nYour question database is ready to use:")
            print(f"  • Excel: {self.excel_path}")
            print(f"  • SQLite: {self.db_path}")
            print(f"\nNext steps:")
            print(f"  1. Create a YAML file with your questions")
            print(f"  2. Run: python main.py your_questions.yaml")
            print(f"  3. Check stats: python main.py --stats")
        else:
            print(f"✗ Setup completed with errors. Please check the messages above.")
        print(f"{'='*60}\n")
    
    def show_stats(self):
        """Show statistics about both Excel and SQLite databases"""
        print(f"\n{'='*60}")
        print(f"Database Statistics")
        print(f"{'='*60}")
        
        # Excel stats
        try:
            wb = self.load_excel()
            ws = wb.active
            excel_total = ws.max_row - 1
            
            # Count by type in Excel
            excel_types = {}
            questions_with_images = 0
            for row in range(2, ws.max_row + 1):
                type_val = ws.cell(row=row, column=11).value  # Type column
                question_image = ws.cell(row=row, column=3).value  # Question image URL
                if type_val:
                    excel_types[type_val] = excel_types.get(type_val, 0) + 1
                if question_image:
                    questions_with_images += 1
            
            print(f"\nExcel Database: {self.excel_path}")
            print(f"  Total Questions: {excel_total}")
            print(f"  Questions with Images: {questions_with_images}")
            print(f"  Breakdown by Type:")
            for type_val, count in sorted(excel_types.items()):
                print(f"    • {type_val}: {count} questions")
        except Exception as e:
            print(f"  ✗ Error reading Excel: {e}")
        
        # SQLite stats
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Total count
            cursor.execute("SELECT COUNT(*) FROM questions")
            db_total = cursor.fetchone()[0]
            
            # Count by type
            cursor.execute("SELECT type, COUNT(*) FROM questions GROUP BY type ORDER BY type")
            db_types = dict(cursor.fetchall())
            
            # Count questions with images
            cursor.execute("SELECT COUNT(*) FROM questions WHERE question_image_url IS NOT NULL AND question_image_url != ''")
            db_images = cursor.fetchone()[0]
            
            print(f"\nSQLite Database: {self.db_path}")
            print(f"  Total Questions: {db_total}")
            print(f"  Questions with Images: {db_images}")
            print(f"  Breakdown by Type:")
            for type_val, count in sorted(db_types.items()):
                print(f"    • {type_val}: {count} questions")
            
            conn.close()
        except Exception as e:
            print(f"  ✗ Error reading SQLite: {e}")
        
        print(f"{'='*60}\n")


def main():
    parser = argparse.ArgumentParser(
        description='Process YAML question files and append to Excel and SQLite databases',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # First time setup - create databases
  python main.py --setup

  # Process single file (interactive)
  python main.py input.yaml

  # Process single file (auto-skip duplicates)
  python main.py input.yaml --skip-duplicates

  # Process all YAML files in a directory
  python main.py --batch ./yaml_files/

  # Show statistics
  python main.py --stats

  # Custom file paths
  python main.py input.yaml --excel custom.xlsx --db custom.db
  
  # Setup with custom paths
  python main.py --setup --excel custom.xlsx --db custom.db
        """
    )
    
    parser.add_argument('input', nargs='?', help='Input YAML file path')
    parser.add_argument('--batch', metavar='DIR', help='Process all YAML files in directory')
    parser.add_argument('--excel', default='questions_database.xlsx', help='Excel database path')
    parser.add_argument('--db', default='questions_database.db', help='SQLite database path')
    parser.add_argument('--skip-duplicates', action='store_true', help='Automatically skip duplicate questions')
    parser.add_argument('--overwrite-duplicates', action='store_true', help='Automatically overwrite duplicate questions')
    parser.add_argument('--setup', action='store_true', help='Initialize database system (create Excel and SQLite databases)')
    parser.add_argument('--create-template', action='store_true', help='Create a new Excel template file (deprecated, use --setup)')
    parser.add_argument('--stats', action='store_true', help='Show database statistics')
    parser.add_argument('--max-tags', type=int, default=4, help='Maximum number of tag columns (default: 4)')
    
    args = parser.parse_args()
    
    # Initialize processor
    processor = QuestionProcessor(
        excel_path=args.excel, 
        db_path=args.db,
        max_tags=args.max_tags,
        auto_init=not args.setup  # Don't auto-init if running setup command
    )
    
    # Handle different modes
    if args.setup:
        processor.setup()
    elif args.create_template:
        print("Note: --create-template is deprecated. Use --setup for full initialization.\n")
        processor.create_excel_template()
    elif args.stats:
        processor.show_stats()
    elif args.batch:
        processor.process_batch(args.batch, args.skip_duplicates, args.overwrite_duplicates)
    elif args.input:
        processor.process_single_file(args.input, args.skip_duplicates, args.overwrite_duplicates)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()